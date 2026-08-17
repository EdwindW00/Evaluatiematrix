"""Excel-export van de (vastgestelde) evaluatiematrix.

MVP-scope: exporteert de matrix zelf (categorieën, criteria, wegingen,
knock-out-markering, bron, toelichting) met werkende SOM-formules voor
sub- en eindtotalen, in Edwins huisstijl. Scoring per leverancier volgt in
een latere iteratie — de kolomindeling is bewust zo opgezet dat leverancier-
kolomblokken er later naast kunnen worden toegevoegd zonder herontwerp.

De structuur is aanpasbaar zodra Edwins eigen bestaande template exact is
vastgesteld (zie evaluatiematrix-app-spec.md).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config import KLEUR_ACCENT, KLEUR_PRIMAIR

PRIMAIR_FILL = PatternFill("solid", fgColor=KLEUR_PRIMAIR)
ACCENT_FILL = PatternFill("solid", fgColor=KLEUR_ACCENT)
LICHT_FILL = PatternFill("solid", fgColor="EAF1FA")
WIT_FILL = PatternFill("solid", fgColor="FFFFFF")
GEEL_WAARSCHUWING = PatternFill("solid", fgColor="FFF3CD")

WITTE_BOLD = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
WITTE_BOLD_TITEL = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
ZWART = Font(color="1A1A1A", size=10, name="Calibri")
ZWART_BOLD = Font(color="1A1A1A", size=10, bold=True, name="Calibri")
ACCENT_BOLD = Font(color=KLEUR_PRIMAIR, size=10, bold=True, name="Calibri")

DUN = Side(style="thin", color="C9D6E5")
RAND = Border(left=DUN, right=DUN, top=DUN, bottom=DUN)

KOLOMMEN = ["Categorie", "Criterium", "Type", "Schaal", "Weging (%)", "Bron", "Toelichting"]


def _header_row(ws: Worksheet, row: int) -> None:
    for i, kop in enumerate(KOLOMMEN, start=1):
        cell = ws.cell(row=row, column=i, value=kop)
        cell.fill = PRIMAIR_FILL
        cell.font = WITTE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = RAND


def export_matrix(project: dict, categorieen: list[dict], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluatiematrix"
    ws.sheet_view.showGridLines = False

    # Titelblok
    ws.merge_cells("A1:G1")
    titel = ws["A1"]
    titel.value = f"Evaluatiematrix — {project.get('naam', '')}"
    titel.font = WITTE_BOLD_TITEL
    titel.fill = PRIMAIR_FILL
    titel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    opdrachtgever = project.get("opdrachtgever") or ""
    sub.value = f"Opdrachtgever: {opdrachtgever}    |    Status matrix: {project.get('matrix_status', 'concept')}"
    sub.font = Font(color=KLEUR_PRIMAIR, italic=True, size=10, name="Calibri")
    ws.row_dimensions[2].height = 18

    header_row = 4
    _header_row(ws, header_row)
    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1
    cat_weging_cellrefs = []
    zebra_on = False

    for cat in categorieen:
        cat_start_row = row
        cat_cell = ws.cell(row=row, column=1, value=cat["naam"])
        cat_cell.font = ACCENT_BOLD
        cat_cell.fill = PatternFill("solid", fgColor="D6E4F5")
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="D6E4F5")
        criteria = cat.get("criteria", [])
        first_crit_row = row + 1

        # Categorie-kop krijgt de SOM van zijn criterium-wegingen (levend, niet bevroren)
        if criteria:
            last_crit_row = first_crit_row + len(criteria) - 1
            ws.cell(row=row, column=5, value=f"=SUM(E{first_crit_row}:E{last_crit_row})")
        else:
            ws.cell(row=row, column=5, value=cat.get("weging", 0))
        ws.cell(row=row, column=5).font = ACCENT_BOLD
        ws.cell(row=row, column=5).number_format = "0"
        cat_weging_cellrefs.append(f"E{cat_start_row}")
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = RAND
        row += 1

        for crit in criteria:
            is_knockout = crit.get("type") == "knock-out"
            fill = LICHT_FILL if zebra_on else WIT_FILL
            zebra_on = not zebra_on

            ws.cell(row=row, column=1, value="")
            ws.cell(row=row, column=2, value=crit["naam"])
            type_label = "Knock-out ☐" if is_knockout else "Gewogen score"
            ws.cell(row=row, column=3, value=type_label)
            ws.cell(row=row, column=4, value="go/no-go" if is_knockout else crit.get("schaal", "0-10"))
            weging_cell = ws.cell(row=row, column=5, value=0 if is_knockout else crit.get("weging", 0))
            weging_cell.number_format = "0"
            ws.cell(row=row, column=6, value=crit.get("bron", ""))
            ws.cell(row=row, column=7, value=crit.get("toelichting", ""))

            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.font = ZWART
                cell.border = RAND
                cell.alignment = Alignment(vertical="top", wrap_text=(col in (6, 7)))
            row += 1

    # Eindtotaal
    total_row = row + 1
    ws.cell(row=total_row, column=1, value="TOTAAL").font = ZWART_BOLD
    total_formula = "=" + "+".join(cat_weging_cellrefs) if cat_weging_cellrefs else "=0"
    total_cell = ws.cell(row=total_row, column=5, value=total_formula)
    total_cell.font = ZWART_BOLD
    total_cell.number_format = "0"
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="F0F0F0")
        ws.cell(row=total_row, column=col).border = RAND

    # Voorwaardelijke opmaak: waarschuwing als totaal niet 100 is
    from openpyxl.formatting.rule import CellIsRule

    ws.conditional_formatting.add(
        total_cell.coordinate,
        CellIsRule(operator="notEqual", formula=["100"], fill=GEEL_WAARSCHUWING),
    )

    if categorieen and not any(c.get("criteria") for c in categorieen):
        pass  # lege matrix; geen extra opmaak nodig

    # Handtekeningvelden
    sig_row = total_row + 3
    ws.merge_cells(f"A{sig_row}:G{sig_row}")
    ws.cell(row=sig_row, column=1, value="Vaststelling").font = ZWART_BOLD
    ws.cell(row=sig_row, column=1).fill = PatternFill("solid", fgColor="D6E4F5")
    for col in range(1, 8):
        ws.cell(row=sig_row, column=col).fill = PatternFill("solid", fgColor="D6E4F5")

    labels = ["Naam", "Functie", "Datum", "Paraaf"]
    for i, label in enumerate(labels):
        r = sig_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = ZWART_BOLD
        ws.merge_cells(f"B{r}:G{r}")
        cell = ws.cell(row=r, column=2, value="")
        cell.border = Border(bottom=Side(style="thin", color="808080"))

    # Kolombreedtes
    widths = [22, 34, 14, 12, 12, 26, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================================
# Fase 8 — volledige export met scores per leverancier
# ============================================================================

LEVERANCIER_KLEUREN = ["4472C4", "2E6DB4", "5B9BD5", "70A9D6", "1A4E8C", "8FBFE0"]
OVERSCHREVEN_FILL = PatternFill("solid", fgColor="FDE68A")
UITGESLOTEN_FILL = PatternFill("solid", fgColor="FFC7CE")
UITGESLOTEN_FONT = Font(color="9C0006", size=10, name="Calibri")
VOLDAAN_FILL = PatternFill("solid", fgColor="D1FAE5")
NIET_VOLDAAN_FILL = PatternFill("solid", fgColor="FEE2E2")
GRIJS_FILL = PatternFill("solid", fgColor="F0F0F0")

FIXED_KOLOMMEN = ["Categorie", "Criterium", "Type", "Schaal", "Weging (%)", "Bron", "Toelichting",
                   "Schaalmin", "Schaalmax"]
N_FIXED = len(FIXED_KOLOMMEN)  # = 9, kolom J (10) is de eerste leverancierkolom


def _parse_schaal(schaal: str | None) -> tuple[float, float]:
    from app.scoring.calculations import parse_schaal
    return parse_schaal(schaal)


def export_full(project: dict, categorieen: list[dict], leveranciers: list[dict],
                 scores_per_leverancier: dict[str, dict], totalen_per_leverancier: dict[str, dict],
                 output_path: Path) -> Path:
    """Volledige export: matrix + score-kolomblok per leverancier, met werkende formules."""
    # Ranking: uitgesloten leveranciers onderaan, verder op eindtotaal
    gerangschikt = sorted(
        leveranciers,
        key=lambda lev: (
            totalen_per_leverancier.get(lev["id"], {}).get("uitgesloten", False),
            -totalen_per_leverancier.get(lev["id"], {}).get("eindtotaal", 0),
        ),
    )
    n_lev = len(gerangschikt)
    n_cols = N_FIXED + n_lev * 3

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluatiematrix"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["H"].hidden = True
    ws.column_dimensions["I"].hidden = True

    # Titelblok
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    titel = ws.cell(row=1, column=1, value=f"Evaluatiematrix — {project.get('naam', '')}")
    titel.font = WITTE_BOLD_TITEL
    titel.fill = PRIMAIR_FILL
    titel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub = ws.cell(row=2, column=1,
                   value=f"Opdrachtgever: {project.get('opdrachtgever') or ''}    |    "
                         f"Status matrix: {project.get('matrix_status', 'concept')}    |    "
                         f"{n_lev} leverancier(s)")
    sub.font = Font(color=KLEUR_PRIMAIR, italic=True, size=10, name="Calibri")

    # ------------------------------------------------------------ ranking-samenvatting
    rank_start = 4
    ws.cell(row=rank_start, column=1, value="Ranking").font = ACCENT_BOLD
    for i, lev in enumerate(gerangschikt):
        totalen = totalen_per_leverancier.get(lev["id"], {})
        r = rank_start + 1 + i
        uitgesloten = totalen.get("uitgesloten", False)
        rangnr = "—" if uitgesloten else str(i + 1)
        status = "UITGESLOTEN (knock-out)" if uitgesloten else "in ranking"
        rang_cell = ws.cell(row=r, column=1, value=rangnr)
        rang_cell.font = ZWART_BOLD
        naam_cell = ws.cell(row=r, column=2, value=lev["naam"])
        eind_cell = ws.cell(row=r, column=3, value=totalen.get("eindtotaal", 0))
        status_cell = ws.cell(row=r, column=4, value=status)
        if uitgesloten:
            for c in (rang_cell, naam_cell, eind_cell, status_cell):
                c.font = UITGESLOTEN_FONT
                c.fill = UITGESLOTEN_FILL

    # ------------------------------------------------------------ hoofdheader
    group_row = rank_start + n_lev + 3
    sub_row = group_row + 1
    ws.merge_cells(start_row=group_row, start_column=1, end_row=group_row, end_column=7)
    gk = ws.cell(row=group_row, column=1, value="Gunningscriteria")
    gk.fill = PRIMAIR_FILL
    gk.font = WITTE_BOLD
    gk.alignment = Alignment(horizontal="center", vertical="center")

    for i, kop in enumerate(FIXED_KOLOMMEN, start=1):
        cell = ws.cell(row=sub_row, column=i, value=kop if i <= 7 else "")
        cell.fill = PRIMAIR_FILL
        cell.font = WITTE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = RAND

    lev_col_start = {}
    for li, lev in enumerate(gerangschikt):
        start_col = N_FIXED + 1 + li * 3
        lev_col_start[lev["id"]] = start_col
        kleur = LEVERANCIER_KLEUREN[li % len(LEVERANCIER_KLEUREN)]
        fill = PatternFill("solid", fgColor=kleur)
        totalen = totalen_per_leverancier.get(lev["id"], {})
        naam_label = lev["naam"] + (" ⚠ UITGESLOTEN" if totalen.get("uitgesloten") else "")
        ws.merge_cells(start_row=group_row, start_column=start_col, end_row=group_row, end_column=start_col + 2)
        gcell = ws.cell(row=group_row, column=start_col, value=naam_label)
        gcell.fill = fill
        gcell.font = WITTE_BOLD
        gcell.alignment = Alignment(horizontal="center", vertical="center")
        for j, sub_kop in enumerate(["Score / Voldaan", "Onderbouwing, citaat & vertrouwen", "Punten (gewogen)"]):
            c = ws.cell(row=sub_row, column=start_col + j, value=sub_kop)
            c.fill = fill
            c.font = WITTE_BOLD
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = RAND

    ws.freeze_panes = ws.cell(row=sub_row + 1, column=N_FIXED + 1).coordinate

    # ------------------------------------------------------------ criteria-rijen
    row = sub_row + 1
    cat_weging_refs = []
    cat_punten_refs = {lev["id"]: [] for lev in gerangschikt}
    zebra = False

    for cat in categorieen:
        cat_row = row
        cat_cell = ws.cell(row=row, column=1, value=cat["naam"])
        cat_cell.font = ACCENT_BOLD
        for col in range(1, n_cols + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="D6E4F5")
            ws.cell(row=row, column=col).border = RAND
        criteria = cat.get("criteria", [])
        first_row = row + 1

        if criteria:
            last_row = first_row + len(criteria) - 1
            ws.cell(row=row, column=5, value=f"=SUM(E{first_row}:E{last_row})")
        else:
            ws.cell(row=row, column=5, value=cat.get("weging", 0))
        ws.cell(row=row, column=5).font = ACCENT_BOLD
        ws.cell(row=row, column=5).number_format = "0"
        cat_weging_refs.append(f"E{cat_row}")

        for lev in gerangschikt:
            start_col = lev_col_start[lev["id"]]
            punten_col_letter = get_column_letter(start_col + 2)
            if criteria:
                last_row = first_row + len(criteria) - 1
                formula = f"=SUM({punten_col_letter}{first_row}:{punten_col_letter}{last_row})"
            else:
                formula = 0
            pcell = ws.cell(row=row, column=start_col + 2, value=formula)
            pcell.font = ACCENT_BOLD
            pcell.number_format = "0.00"
            cat_punten_refs[lev["id"]].append(f"{punten_col_letter}{cat_row}")
        row += 1

        for crit in criteria:
            is_ko = crit.get("type") == "knock-out"
            fill = LICHT_FILL if zebra else WIT_FILL
            zebra = not zebra
            lo, hi = _parse_schaal(crit.get("schaal"))

            ws.cell(row=row, column=2, value=crit["naam"])
            ws.cell(row=row, column=3, value="Knock-out" if is_ko else "Gewogen score")
            ws.cell(row=row, column=4, value="go/no-go" if is_ko else crit.get("schaal", "0-10"))
            wcell = ws.cell(row=row, column=5, value=0 if is_ko else crit.get("weging", 0))
            wcell.number_format = "0"
            ws.cell(row=row, column=6, value=crit.get("bron", ""))
            ws.cell(row=row, column=7, value=crit.get("toelichting", ""))
            ws.cell(row=row, column=8, value=lo)
            ws.cell(row=row, column=9, value=hi)

            for col in range(1, N_FIXED + 1):
                c = ws.cell(row=row, column=col)
                c.fill = fill
                c.font = ZWART
                c.border = RAND
                c.alignment = Alignment(vertical="top", wrap_text=(col in (6, 7)))

            for lev in gerangschikt:
                start_col = lev_col_start[lev["id"]]
                score_col, ond_col, punten_col = start_col, start_col + 1, start_col + 2
                score_letter = get_column_letter(score_col)
                min_letter, max_letter, weging_letter = "H", "I", "E"
                rec = scores_per_leverancier.get(lev["id"], {}).get(crit["id"])
                overschreven = bool(rec and rec.get("overschreven_door_gebruiker"))

                if is_ko:
                    voldaan = rec.get("voldaan") if rec else None
                    label = {1: "☑ Voldaan", 0: "☐ Niet voldaan"}.get(voldaan, "☐ Onbekend")
                    scell = ws.cell(row=row, column=score_col, value=label)
                    if voldaan == 1:
                        scell.fill = VOLDAAN_FILL
                    elif voldaan == 0:
                        scell.fill = NIET_VOLDAAN_FILL
                    else:
                        scell.fill = fill
                    ws.cell(row=row, column=punten_col, value="n.v.t.")
                else:
                    score_val = rec.get("score") if rec else None
                    scell = ws.cell(row=row, column=score_col, value=score_val)
                    scell.fill = OVERSCHREVEN_FILL if overschreven else fill
                    scell.number_format = "0.0"
                    pcell = ws.cell(
                        row=row, column=punten_col,
                        value=(f"=IFERROR(({score_letter}{row}-{min_letter}{row})/"
                               f"({max_letter}{row}-{min_letter}{row})*{weging_letter}{row},0)"),
                    )
                    pcell.number_format = "0.00"
                    pcell.fill = fill

                onderbouwing_delen = []
                if rec:
                    if rec.get("onderbouwing"):
                        onderbouwing_delen.append(rec["onderbouwing"])
                    if rec.get("citaat"):
                        onderbouwing_delen.append(f'"{rec["citaat"]}"')
                    if rec.get("vertrouwen"):
                        onderbouwing_delen.append(f'[{rec["vertrouwen"]}]')
                    if rec.get("gebruiker_commentaar"):
                        onderbouwing_delen.append(f'Commentaar Edwin: {rec["gebruiker_commentaar"]}')
                    if rec.get("nader_verifieren"):
                        onderbouwing_delen.append("⚠ nader te verifiëren met leverancier")
                ocell = ws.cell(row=row, column=ond_col, value="\n".join(onderbouwing_delen))
                ocell.fill = fill
                ocell.font = ZWART
                ocell.alignment = Alignment(vertical="top", wrap_text=True)
                ocell.border = RAND

                scell.font = ZWART_BOLD if overschreven else ZWART
                scell.border = RAND
                scell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
            row += 1

    # ------------------------------------------------------------ eindtotaalrij
    total_row = row + 1
    ws.cell(row=total_row, column=1, value="TOTAAL").font = ZWART_BOLD
    totaal_weging_formula = "=" + "+".join(cat_weging_refs) if cat_weging_refs else "=0"
    tcell = ws.cell(row=total_row, column=5, value=totaal_weging_formula)
    tcell.font = ZWART_BOLD
    tcell.number_format = "0"
    for col in range(1, N_FIXED + 1):
        ws.cell(row=total_row, column=col).fill = GRIJS_FILL
        ws.cell(row=total_row, column=col).border = RAND

    for lev in gerangschikt:
        start_col = lev_col_start[lev["id"]]
        punten_col = start_col + 2
        refs = cat_punten_refs[lev["id"]]
        formula = "=" + "+".join(refs) if refs else "=0"
        cell = ws.cell(row=total_row, column=punten_col, value=formula)
        cell.font = ZWART_BOLD
        cell.number_format = "0.00"
        cell.fill = GRIJS_FILL
        for col in (start_col, start_col + 1):
            ws.cell(row=total_row, column=col).fill = GRIJS_FILL

    # ------------------------------------------------------------ handtekeningvelden
    sig_row = total_row + 3
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=n_cols)
    ws.cell(row=sig_row, column=1, value="Vaststelling").font = ZWART_BOLD
    for col in range(1, n_cols + 1):
        ws.cell(row=sig_row, column=col).fill = PatternFill("solid", fgColor="D6E4F5")

    for i, label in enumerate(["Naam", "Functie", "Datum", "Paraaf"]):
        r = sig_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = ZWART_BOLD
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=min(7, n_cols))
        ws.cell(row=r, column=2).border = Border(bottom=Side(style="thin", color="808080"))

    # ------------------------------------------------------------ kolombreedtes
    widths = [20, 30, 12, 10, 10, 22, 32, 6, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for li in range(n_lev):
        start_col = N_FIXED + 1 + li * 3
        ws.column_dimensions[get_column_letter(start_col)].width = 16
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 42
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 14

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
